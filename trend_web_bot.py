import os
import builtins
import re
import json
import sys

import datetime
import time

import pandas as pd
from openpyxl import load_workbook

import requests
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from bs4 import BeautifulSoup

import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import ttk


#____________________________________________
#Constants

EXCEL_FILE = 'OS_full_list.xlsx'
SHEET_NAME = 'Trend_OS_full_list'
OUTPUT = 'scan_results_'+str(datetime.datetime.now()).split('.')[0].replace(' ','').replace(':','').replace('-','')+'.xlsx'
LOGIN_FILE = 'OS_logins.xlsx'
PARAMETER_TABLE_LIST = {'id':'parameterTable', 'name':'Adjust'}
LOADTIME = 1.8  #used for wait time before scanning/testing page
LOGFILE = "trend_web_bot.log"
ERRORLOG = "error.log"

#______________________________________________


#functions

def find_origVal(soup, target_text):
    #used for editable fields
    log.login(f"find_origVal fct, looking for {target_text}")
    try:
        val = find_associated_element(soup, target_text)
        log.login(f"looking in {val}")
        origVal=val.find('input', attrs={"name":re.compile("origVal$")})['value']
        log.login(f"Found origVal : {origVal}")
        return origVal
    except Exception as e:
        log.login(f"Could not find origVal - {e}")

def write_newVal(html_content, target_text, value, driver):

    def write_newVal_editable(html_content, target_text, value, driver):
        #not to be used directly, implemented through write_newVal
        log.login(f"write_newVal_editable fct, looking for {target_text}")
        try:
            log.login(f"newVal id: {target_text}")
            newVal = driver.find_element(By.ID, target_text)
            log.login(f"newVal: {newVal}")
            newVal.click()
            newVal.clear()
            newVal.send_keys(value)
            return True
        except Exception as e:
            log.login(f"write_newVal_editable fct error: {e}")
            return False
        
    def write_newVal_ass(html_content, target_text, value, driver):
        #not to be used directly, implemented through write_newVal
        log.login(f"write_newVal_ass fct, looking for {target_text}")
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            val = find_associated_element(soup, target_text)
            log.login(f"looking in: {val}")
            newVal_el=val.find('input', attrs={"name":re.compile("newVal$")})
            log.login(f"Found newVal el: {newVal_el}")
            newVal_id = newVal_el['id']
            log.login(f"newVal id: {newVal_id}")
            newVal = driver.find_element("id", newVal_id)
            log.login(f"newVal: {newVal}")
            newVal.clear()
            newVal.send_keys(value)
            return True
        except Exception as e:
            log.login(f"write_newVal_ass fct error: {e}")
            return False
        
    log.login(f"write_newVal fct, looking for {target_text}")
    test = write_newVal_editable(html_content, target_text, value, driver)
    log.login(f"tried to find editable field. Success = {test}")
    if test == False:
        test = write_newVal_ass(html_content, target_text, value, driver)
        log.login(f"tried to find non-editable, associated field. Success = {test}")
    log.login(f"write_newVal fct completed. Success = {test}")
    return test
  
def select_newVal(html_content, target_text, value, driver):
    log.login(f"select_newVal fct, looking for {target_text}")
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        val = find_associated_element(soup, target_text)
        log.login(f"looking in {val}")
        newVal_el=val.find('select', attrs={"name":re.compile("newVal$")})
        log.login(f"Found newVal el: {newVal_el}")
        newVal_id = newVal_el['id']
        log.login(f"newVal id: {newVal_id}")
        newVal = driver.find_element("id", newVal_id)
        log.login(f"newVal: {newVal}")
        newVal=Select(newVal)
        newVal.select_by_visible_text(value)
        return True
    except Exception as e:
        log.login(f"Could not select newVal - {e}")
        return False
    
def submit(driver):
    log.login(f"submit fct")
    try:
        submit_button = driver.find_element("xpath", '//input[@type="image" and @src="images/send.gif"]')
        submit_button.click()
        return True
    except Exception as e:
        log.login(f"Could not submit - {e}")
        return False

def find_associated_element(soup, target_text):
    #used for non editable fields
    log.login(f"find_associated_element fct, looking for {target_text}")
    #log.login(f"in {soup}")
    try:
        # Find all <td> elements with class="pName"
        p_name_elements = soup.find_all('td', class_='pName')
        log.login(f"p_name_elements: {p_name_elements}")
        
        # Iterate over the <td> elements with class="pName"
        for p_name_element in p_name_elements:
            # Check if the text content matches the target_text
            if p_name_element.get_text(strip=True) == target_text:
                # Get the next sibling <td> element with class="pValue"
                p_value_element = p_name_element.find_next_sibling('td', class_='pValue')
                if p_value_element:
                    log.login(f"find_associated_element success, returning {p_value_element}")
                    return p_value_element
    except Exception as e:
        log.login(f"Could not find associated element - {e}")

def visit_webpage_selenium(url, driver):
    log.login("visit_webpage_selenium fct")
    try:
        # Navigate to the webpage
        driver.get(f"http://{url}")
        page_source = driver.page_source

        # Check if the page loaded successfully
        if ("404 Not Found" not in page_source and page_source != None and page_source !=""):
            log.login(f"Successfully visited {url}")
            print("Content:")
            print(page_source)
            return page_source, True
        else:
            log.login(f"Failed to visit {url}. Page not found.")
            return "no answer", False
    except TimeoutException as e:
        log.login(f"Timeout visiting {url}: {e}")
        return "timeout error", False
    
    except WebDriverException as e:
            log.login(f"Web driver exception: {url}: {e}")
            return "access error", False
    
def open_xls(xls, sheetname = None):
    log.login("opening : "+str(xls))
    try:
        data = pd.read_excel(xls, SHEET_NAME=sheetname)
        return data
    except Exception as e:
        log.login(f"excel sheet not available: {e}")

def get_column_number(sheet, target_value):
    log.login(f"get_column_number function, target value: {target_value} in sheet: {sheet}")
    # Iterate over cells in the first row of the sheet
    result = (-1)
    try:
        for cell in sheet[1]:
            # Check if the cell value matches the target value
            if cell.value == target_value:
                # Return the column number (index) of the matching cell
                result = cell.column
                log.login(f"column found: {result}")
        log.login(f"get_column_number function failed, column not found")   
    except Exception as e:
        log.login(f"get_column_number function error : {e}")
    finally:
        log.login(f"get_column_number completed result = {result}")
        return result

def get_maximum_cols(xls, sheet, row_index=1):
    #not in use, same result as max_col
    log.login(f"get_maximum_cols function init, row: {row_index} in sheet: {sheet}")
    cols = 0
    if (cell.value is not None for cell in xls[sheet][row_index]):
        cols += 1
    log.login(f"get_maximum_cols function completes, cols: {cols}")
    return cols

def update_xls_prop_sheet(controller, scrape_res, xls, index):
    log.login(f"update_xls_prop_sheet function")
    #if subpages
    try:
        log.login(f"property : {scrape_res}")
        page = scrape_res[0]
        props = scrape_res[1]
        log.login(f"page : {page}, props : {props}, index : {index}")
        OS_info = {'site':controller.site,'Lan':controller.lan,'OS':controller.os,'IP':controller.ip}
        OS_info.update(props)
        props = OS_info.copy()
        #sheet = subpage["Type"]
        if page not in xls.sheetnames:
            xls.create_sheet(page)
        log.login(f"sheet : {page}")
        for property in props:
            column_nb = get_column_number(xls[page], property)
            log.login(f"for property {property}, value is {props[property]}")
            if column_nb < 0:
                column_nb = len(xls[page][1])+1
                log.login(f"property {property} not found, adding to sheet {page} at column {column_nb}")
                xls[page].cell(row=1, column=(column_nb)).value = property
            xls[page].cell(row=index, column=column_nb).value = props[property]

            log.login(f"update_xls_prop_sheet function complete")
    except Exception as e:
        log.login(f"update_xls_prop_sheet function error : {e}")

#legacy
def clean_prop_name(prop):
    log.login(f"clean_prop_name fct init, prop: {prop}")
    try:
        result = re.sub("[{}]","",prop)
        log.login(f"clean_prop_name fct result: {result}")
        return result
    except Exception as e:
        log.login(f"clean_prop_name fct error : {e}")

def fetch_users(logins=LOGIN_FILE):
    #return user, password, pin in strings
    log.login(f"fetch_users function")
    try:
        login_list = load_workbook(logins, read_only = True)
        login_sheet = login_list.worksheets[0]
        #log.login(f"fetch_users  - sheet: {login_sheet}")
        user_col = get_column_number(login_sheet,'user')-1
        pw_col = get_column_number(login_sheet,'password')-1
        pin_col = get_column_number(login_sheet,'pin')-1
        #log.login(f"fetch_columns success - user: {user_col} password: {pw_col} pin: {pin_col}")
        user_list=[]
        for row in login_sheet.iter_rows(min_row=2):
            log.login(f"row: {row}")
            user = []
            user.append(row[user_col])
            user.append(row[pw_col])
            user.append(row[pin_col])
            user_list.append(user)
            #log.login(f"fetch_users list update success:  {user}")
        log.login(f"fetch_users list completed:  {user_list}")
        return user_list
    except Exception as e:
        log.login(f"fetch_users function error : {e}")
        return [["","",""]]

def check_access(page_source):
    log.login(f"check_access function")
    try:
        errlist=("Error 403 Forbidden","Error 404","Session Expired")
        test = False
        for err in errlist:
            if (err not in page_source):
                test = True
        log.login(f"check_access function done, success: {test}")
        return test
    except Exception as e:
        log.login(f"check_access function error : {e}")

class scraper:
    def __init__(self,controller,driver,pages_list, param_list, user_list, output, index):
        log.login(f"scraper object init with pages: {pages_list} and param: {param_list}")
        self.OS = controller
        self.url = controller.ip
        self.driver = driver
        self.pages_list = pages_list
        self.param_list = param_list
        self.user_list = user_list
        self.output = output
        self.scrape_index = index
        self.wait = WebDriverWait(self.driver, LOADTIME)  # Wait for a maximum time, in seconds
        for param in self.param_list:
            log.login(f"for param {param}, value is {param_list[param]}")
        log.login(f"scraper object init done")

    def scrape_all(self, pages_list):
        log.login(f"scrape_all fct - properties: {pages_list}")
        try:
            if self.ctlr_online_test():
                self.ctlr_login(self.user_list)
                with open('json_dat/pages.json', 'r') as json_file:
                    pages = json.load(json_file)
                    log.login(f"pages json loaded: {pages}")
                    for page, link in pages.items():
                        if page in pages_list:
                            log.login(f"scraping page: {page}")
                            success = self.scrape_page(page, 0, 2)

                    return  True
            else:
                    return [], False
        except Exception as e:
            log.login(f"scrape_all error: {e}")
            return  False

    def scrape_page(self, page, count, max_count, path = ""):
        log.login(f"scrape_page fct: {page}")
        if path == "":
            path = page
        else:
            path = f"{path}_{page}"
        try:
            page_src = self.driver.page_source
            self.open_link_same_tab(By.XPATH,page)

            try:
                html_content = self.driver.page_source
                parameter_table, parameters = self.find_param_table(html_content)
            
                all_fields = dict()
                for parameter in parameters:
                    log.login(f"parameter {parameters.index(parameter)}: {parameter}")
                    field = parameter.string
                    value, success = self.scrape_element(parameter_table, parameter.string)
                    log.login(f"parameter {parameters.index(parameter)}: {field} = {value}")
                    all_fields.update({field : value})
            
                if success:
                    update_xls_prop_sheet(self.OS, [path,all_fields], self.output, self.scrape_index)
                    self.driver.back()
                log.login(f"scrape_page fct complete, values: {[path,all_fields]}")
                return success
            except Exception as e:
                log.login(f"scrape_page error: {e}")
                if count < max_count:
                    log.login(f"scrape_page one level deeper")
                    count += 1
                    links, success = self.get_links()
                    for sub in links:
                        log.login(f"sublink is: {sub}")
                        success = self.scrape_page(sub.text, count, max_count, path)
                        log.login(f"for sublink {sub}, success = {success}")
                        success *= success
                        log.login(f"after sublink {sub}, success = {success}")
                    log.login(f"success: {success}")
                    return success
                else:
                    for _ in range(max_count):
                        self.driver.back()
                    return  False   
        except Exception as e:
            log.login(f"scrape_page error - could not access page : {e}")
            
    def scrape_element(self, soup, element):
        log.login("scrape_element fct")
        value = ""
        success = True
        try:
            value = find_origVal(soup, element)
        except Exception as e:
            log.login(f"scrape_element error for origVal: {e} - this can be normal if not editable field")
        if value == None:
            try:
                value = find_associated_element(soup, element).get_text(strip=True)
                log.login(f"scrape_element success, non editable field")
            except Exception as e:
                log.login(f"scrape_element error for associated_element: {e}")
                value = "error"
                success = False
        log.login(f"scrape_element fct done, value: {value}, success: {success}")
        return value, success
    
    def find_custom_selector(self, params, soup):
        #not working
        log.login(f"find_custom_selector init with {params}")
        tags = []
        for param in params:
            log.login(f"for param {param}, value is {params[param]}")
            tag = BeautifulSoup(f"('{param}')=='{params[param]}'")
            tags.append(tag)
        log.login(f"tags to look for: {tags}")
        tag = soup.get(field for field in tags)
        result = soup.find(tag)
        log.login(f"find_custom_selector fct ended, result: {result}")    
        return result

    def find_param_table(self,html_page):
        try:
            log.login(f"find_param_table fct init")
            soup = BeautifulSoup(html_page, 'html.parser')
            log.login("content cast into soup")
            #log.login(f"soup is: {soup}")
            try:
                parameter_table = self.find_custom_selector(self.param_list,soup)
                #parameter_table = soup.find(lambda tag: tag.get('id')=='parameterTable' or tag.get('name')=='Adjust')
                log.login(f"parameter table is: {parameter_table}")
                parameters = parameter_table.find_all('td', class_='pName')
            except Exception as e:
                log.login(f"Could not find subdividers: {e}")
                paremeter_table = e
                parameters = ""
            log.login(f"find_param_table fct done, parameter_table: {parameters}")
            return parameter_table, parameters
        except Exception as e:
            log.login(f"find_param_table fct error: {e}")

    def get_links(self):
        log.login("get_links fct")
        try:
            #html_content = visit_webpage_selenium(f"{url}", self.driver)
            page_src = self.driver.page_source
            #page_link = self.driver.find_element("link text", page)
            #page_link.click()
            #if (html_content[1] == False) : return "visit error", False
            soup = BeautifulSoup(page_src, 'html.parser')
            log.login("content cast into soup")
            try:
                mainContent = soup.find(id=['mainContent','maindata'])
            except Exception as e:
                    log.login(f"Could not find subdivider: {e}")
            ablocks = mainContent.find_all('a')
            log.login(f"links: {ablocks}")
            return ablocks, True
        except Exception as e:
            log.login(f"get_links error: {e}")
            return "get_links error", False

    def get_subpages(self):
        log.login("get_subpages fct")
        try:
            html_content = self.driver.page_source
            #if (html_content[1] == False) : return "visit error", False
            soup = BeautifulSoup(html_content, 'html.parser')
            log.login("content cast into soup")
            try:
                mainContent = soup.find(id=['mainContent','maindata'])
            except Exception as e:
                    log.login(f"Could not find subdivider: {e}")
            ablocks = mainContent.find_all('a')
            #links = []
            #for link in ablocks:
            #    links.append(link.get('href'))
            log.login(f"links: {ablocks}")
            return ablocks, True
        except Exception as e:
            log.login(f"get_subpages error: {e}")
            return "get_subpages error", False
        
    def get_all_pages(self):
        log.login("get_all_pages fct")
        try:
            html_content = visit_webpage_selenium(f"{self.url}/modules.htm", self.driver)  # Replace with the IP address you want to visit
            if (html_content[1] == False) : return "visit error", False
            soup = BeautifulSoup(html_content[0], 'html.parser')
            log.login("content cast into soup")
            try:
                mainTable = soup.find(class_=('sideMenu'))
                log.login(f"mainTable: {mainTable}")
            except Exception as e:
                    log.login(f"Could not find subdivider: {e}")
            links = mainTable.find_all('a')
            log.login(f"links: {links}")
            pages = dict()
            for link in links:
                log.login(f"link {links.index(link)}: {link}")
                pages.update({link.string : link.get('href')})
            log.login(f"pages :{pages}")
            return pages, True
        except Exception as e:
            log.login(f"get_all_pages error: Error getting controller pages: {e}")
            return "get_all_pages error", False

    def ctlr_login(self,user_list):
        log.login(f"ctlr_login fct")
        try:
            for user in user_list:
                if (self.ctlr_access_test()==False):
                    log.login(f"access denied, login attempt")
                    html_content = visit_webpage_selenium(f"{self.url}/login.htm", self.driver)
                    if (html_content[1] == False) : 
                        log.login(f"error: login page access denied")
                        return False
                    else:
                        log.login(f"login page accessed succesfully")
                        #log.login(f"page source: {html_content[0]}")
                        write_newVal(html_content[0], "username", user[0].value, self.driver)
                        write_newVal(html_content[0], "password", user[1].value, self.driver)
                        login_button = self.driver.find_element("xpath", '//input[@type="submit" and @value="Login"]')
                        login_button.click()
                else:
                    log.login(f"ctlr_login fct completed: access granted")
                    return True
            log.login(f"ctlr_login fct completed: access denied")
        except Exception as e:
            log.login(f"ctlr_login error: Error login into controller: {e}")
            return "ctlr_login error", False

    def ctlr_access_test(self):
        log.login(f"ctlr_access_test fct")
        try:
            #page_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, 'MODULES')))
            page_link = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@class='mainlink' and contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'modules')]")))
            log.login(f"page_link is: {page_link}")
            page_link.click()
            #page_link = wait.until(EC.visibility_of_element_located((By.LINK_TEXT, "Networks")))
            page_link = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@class='submenu' and contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'),'networks')]")))
            page_link.click()
            test = True
            self.driver.back()
            log.login(f"ctlr_access_test fct completed. Result: {test}")
            return test
        except Exception as e:
            log.login(f"ctlr_access_test error: {e}")
            log.login(f"page_source is: {self.driver.page_source}")
            log.login(f"ctlr_access_test fct completed. Result: False")
            return False   

    def ctlr_online_test(self):
        log.login(f"ctlr_online_test fct")
        try:
            html_content = visit_webpage_selenium(self.url, self.driver)
            test_rslt = html_content[1]
            log.login(f"ctlr_online_test fct completed : {test_rslt}")    
            return test_rslt
        except Exception as e:
            log.login(f"ctlr_online_test fct error: {e}")
            return False

    def open_link_same_tab(self,by,value):
            log.login(f"open_link_same_tab fct")
            try:
                page_link = self.wait.until(EC.element_to_be_clickable((by, f"//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{value.lower()}')]")))
                self.driver.execute_script("arguments[0].setAttribute('target', '_self');",page_link)
                page_link.click()
                log.login(f"open_link_same_tab fct completed")
                return True
            except Exception as e:
                log.login(f"open_link_same_tab fct error: {e}")
                return False
    
class controller:
    def __init__(self, row):
        self.site = row['siteLabel']
        self.lan = row['LanNo']
        self.os = row['NodeAddress']
        self.ip = row['nodeIpAddr']

class logger:
    def __init__(self, filename):
        self.filename = filename
        self.fp=None   

    def __enter__(self):
        print("__enter__")
        self.fp=open(self.filename,"w")
        self.login("\nnew log started\n************************************\n")
        return self   
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        print("__exit__")
        self.fp.close()


    def login(self,text):
        self.fp.writelines(str(datetime.datetime.now())+" : "+str(text)+"\n")
        print(str(text))

#____________________________________________
#GUI functions
class replace_interface(tk.Frame):

    def prop_drop_down(self, event):
        #make property list
        try:
            with open(f"json_dat/{self.page.get()}.json", 'r') as json_file:
                pages = json.load(json_file)
                property_list = []
                for page in pages:
                    for prop in page:
                        property_list.append(prop)
                sorted(set(property_list))
                log.login(f"property_list : {property_list}")
                self.prop["values"]=property_list

        except Exception as e:
            log.login(f"error creating property_list: {e}")

    def on_checkbox_toggle(self):
            global Replace 
            Replace = not Replace
            log.login(f"Replace mode = {Replace}")
            return Replace
    
    def __init__(self, parent, pages):
        tk.Frame.__init__(self, parent)
        buffer = tk.Frame(self, width=20, height=20)

        #create the frames
        self.original_fr = tk.Frame(self, height = 10, width = 450)
        self.new_fr = tk.Frame(self, height = 10, width = 450)
        self.prop_fr = tk.Frame(self, height = 10, width = 120)
        self.filter_fr = tk.Frame(self, height = 10, width = 120)
        buffer.pack(expand=True)
        self.original_fr.pack(expand=True)
        buffer.pack(expand=True)
        self.new_fr.pack(expand=True)
        buffer.pack(expand=True)
        self.prop_fr.pack(expand=True)
        self.filter_fr.pack(expand=True)

        #create the replacing interface
        old_label = tk.Label(self, text="Replace this:")
        self.replace_this = tk.Entry(self, width = 20)
        new_label = tk.Label(self, text="By this:")
        self.by_this = tk.Entry(self, width = 20)
        old_label.pack(in_=self.original_fr, fill=tk.BOTH, side=tk.LEFT)
        self.replace_this.pack(in_=self.original_fr, fill=tk.BOTH, side=tk.RIGHT)
        new_label.pack(in_=self.new_fr, fill=tk.BOTH, side=tk.LEFT)
        self.by_this.pack(in_=self.new_fr, fill=tk.BOTH, side=tk.RIGHT)

        page_label = tk.Label(self, text="Page:")
        self.page = ttk.Combobox(state="readonly", values=pages)
        self.page.bind("<<ComboboxSelected>>", self.prop_drop_down)
        property_label = tk.Label(self, text="Property:")
        self.prop = ttk.Combobox(values="")
        page_label.pack(in_=self.prop_fr, side=tk.LEFT)
        self.page.pack(in_=self.prop_fr, side=tk.LEFT)
        tk.Frame(self, width=40, height=10).pack(in_=self.prop_fr, side=tk.LEFT)
        property_label.pack(in_=self.prop_fr, side=tk.LEFT)
        self.prop.pack(in_=self.prop_fr, side=tk.LEFT)

        #replace checkbox
        r = tk.Checkbutton(self, text="Replace", variable=Replace, width=10, height=2, command=self.on_checkbox_toggle)
        r.pack(anchor = "center")

        buffer.pack(side=tk.TOP)

class GUI:
    def cancel(self):
        sys.exit()

    def replace(self, scraper, page, prop, old, new, count=0, max_count=2):
            log.login(f"replace fct: for property: {page}/{prop}, replace {old} by {new}")
            try:
                #find right page
                scraper.open_link_same_tab(By.XPATH,page)

                try:
                    #scrape element
                    page_src = scraper.driver.page_source
                    parameter_table, parameters = scraper.find_param_table(page_src)
                    value, success = scraper.scrape_element(parameter_table, prop)
                    log.login(f"found property: {prop} in page :{page}, value is: {value}")
                    if success is not True:
                        raise Exception ("could not find value")

                    if (value == old):
                        log.login(f"old value {old} is to be replaced by {new}")
                        try:
                            select_newVal(page_src, prop, new, scraper.driver)
                        except Exception as e:
                            log.login(f"new value could not be selected: {e}")      
                        try:
                            write_newVal(page_src, prop, new, scraper.driver)
                        except Exception as e:
                            log.login(f"new value could not be written: {e}")     
                        try:
                            submit(scraper.driver)
                        except Exception as e:
                            log.login(f"could not submit: {e}")
                    else:
                          log.login(f"old value {old} not found, nothing replaced")
                    
                    if success:
                        scraper.driver.back()
                    return success    

                except Exception as e:
                            log.login(f"replace error: {e}")
                            if count < max_count:
                                log.login(f"replace one level deeper")
                                count += 1
                                links, success = scraper.get_links()
                                for sub in links:
                                    log.login(f"sublink is: {sub}")
                                    success = self.replace(scraper, sub.text, prop, old, new, count)
                                    log.login(f"for sublink {sub}, success = {success}")
                                    success *= success
                                    log.login(f"after sublink {sub}, success = {success}")
                                log.login(f"success: {success}")
                                return success
                            else:
                                for _ in range(max_count):
                                    scraper.driver.back()
            except Exception as e:
                log.login(f"replace fct error: {e}")
                return False   
            
    def scan(self):
        #global Replace 
        Confirm = False

        def execute(self):

            log.login(f"sites to action: {self.selected_sites}")
            log.login(f"pages to read: {self.selected_properties}")

            #init output
            excel_list = load_workbook(EXCEL_FILE)

            # Initialize the WebDriver (replace 'chromedriver' with the path to your driver executable)
            #options = webdriver.ChromeOptions()
            options = webdriver.EdgeOptions()
            options.add_argument('ignore-certificate-errors')
            options.add_argument('acceptInsecureCerts')
            #with webdriver.Chrome(options=options) as driver:
            with webdriver.Edge(options=options) as driver:
                try:
                    # Loop through all rows using iterrows()
                    scrape_index = 1
                    for index, row in os_list.iterrows():
                        try:
                            log.login("new row of excel sheet")
                            TrendCont = controller(row)
                            contrLogLine = f"{TrendCont.site} - L{TrendCont.lan:n}OS{TrendCont.os:n} - {TrendCont.ip}"
                            if TrendCont.site in self.selected_sites and (TrendCont.os != 126):
                                log.login(f"controller to check: {contrLogLine}")
                                if (TrendCont.ip == "#N/A#" or TrendCont.ip == "" or TrendCont.ip == "inv" or pd.isna(TrendCont.ip)):
                                    log.login(f"controller not visitable : {contrLogLine}")
                                else :
                                    scrape_index += 1
                                    os_scraper = scraper(TrendCont,driver,self.selected_properties, PARAMETER_TABLE_LIST, self.user_list,excel_list, scrape_index)
                                    success = os_scraper.scrape_all(os_scraper.pages_list)
                                    #replace function below
                                    if Replace == True:
                                        success = self.replace(os_scraper, self.replace_bar.page.get(), self.replace_bar.prop.get(), self.replace_bar.replace_this.get(), self.replace_bar.by_this.get())
                        except Exception as e:
                            log.login(f"Controller failure, skipping controller {row} - "+str(e))
                            error.login(f"{contrLogLine} could not be accessed")
                    try:
                        #add in code for custom output in case of failure
                        excel_list.save(OUTPUT)
                    except Exception as e:
                        log.login(f"Write error: {e}")
                except Exception as e:
                    log.login("Major failure, exiting now - "+str(e))
                driver.close()
                log.login("Done")

        def confirm_replace():
            nonlocal popup
            Confirm = True
            popup.destroy
            execute(self)

        if Confirm == False and Replace == True :
            popup = tk.Toplevel(self.root)
            popup.title("Replace")

            confirm_button = tk.Button(popup, text="Confirm", command=confirm_replace)
            confirm_button.pack()
            cancel_button = tk.Button(popup, text="Cancel", command=self.cancel)
            cancel_button.pack()
            confirm_text = tk.Label(popup, text = "Are you sure you want to write to controllers?",wraplength=150, width=35, height=15)
            confirm_text.pack()

        else:
            execute(self)

    def createChkbx(self):
        return GUI.checkbox_list(self)
    
    class checkbox_list:

        def __init__(self, outer_instance):
            self.outer_instance = outer_instance
            self.checkboxes = []
            self.selected_values = []

        def show_checkbox_list(self, ckb_list):
            
            def confirm_selection(self):
                self.selected_values = [item for item, var in self.checkboxes if var.get()]
                popup.destroy()

            def toggle_select_all(self):
                select_all_state = select_all_var.get()
                for var in self.checkboxes:
                    #login(f"var is: {var[1]}")
                    var[1].set(select_all_state)
            
            popup = tk.Toplevel(self.outer_instance.root)
            popup.title("Checkbox List")
            

            # Create a variable for "Select All" checkbox
            select_all_var = tk.BooleanVar()
            select_all_var.set(False)  # Initially not selected

            # Calculate number of columns based on the number of options
            num_columns = 3
            num_options = len(ckb_list)+1
            num_rows = -(-num_options // num_columns)  # Equivalent to math.ceil(num_options / num_columns)

            # Create the "Select All" checkbox
            select_all_checkbox = tk.Checkbutton(popup, text="Select All", variable=select_all_var, command=lambda: toggle_select_all(self))
            select_all_checkbox.grid(row=0, column=0, sticky="w")

            for i, item in enumerate(ckb_list):
                row = (i+1) // num_columns
                column = (i+1) % num_columns
                var = tk.BooleanVar()
                checkbtn = tk.Checkbutton(popup, text=item, variable=var)
                checkbtn.grid(row=row, column=column, sticky="w")
                self.checkboxes.append((item, var))

            confirm_button = tk.Button(popup, text="Confirm", command=lambda: confirm_selection(self))
            confirm_button.grid(row=num_rows, columnspan=num_columns, pady=10)

            popup.grab_set()  # Make the popup modal
            popup.wait_window()  # Wait for the popup window to close
            log.login(f"selected values : {self.selected_values}")
            
            return self.selected_values

    def select_property(self, property_list):
       property_chkbx = self.checkbox_list(self)
       log.login(f"self.selected_properties: {self.selected_properties}")
       self.selected_properties = property_chkbx.show_checkbox_list(property_list)
       log.login(f"self.selected_properties: {self.selected_properties}")

    def select_sites(self, sites_list):
        sites_chkbx = self.checkbox_list(self)
        log.login(f"self.selected_sites: {self.selected_sites}")
        self.selected_sites = sites_chkbx.show_checkbox_list(sites_list)
        log.login(f"self.selected_sites: {self.selected_sites}")
    
    def toggle_widget(self,widget):
        if widget.winfo_ismapped():
            widget.pack_forget()
        else:
            widget.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    def __init__(self, sites_list, property_list):
        global Replace
        self.sites_list = sites_list
        self.property_list = property_list
        self.selected_sites = []
        self.selected_properties = []
                    
        #fetch users once
        self.user_list = fetch_users()

        #Tkinter GUI
        root = tk.Tk()
        root.title("Trend Alarm destination crawler")
        root.minsize(920,200)
        #root.geometry("920x300")

        
        # create the main sections of the layout, 
        # and lay them out
        self.replace_bar = replace_interface(root,property_list)
        buffer = tk.Frame(root, width=200, height=20)
        top = tk.Frame(root)
        middle = tk.Frame(root)
        bottom = tk.Frame(root)
        buffer.pack(side=tk.TOP)
        top.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        middle.pack(anchor='center')
        bottom.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
        buffer.pack(side=tk.BOTTOM)
       


        # create the widgets for the top part of the GUI,
        # and lay them out
        s = tk.Button(root, text="Select sites", width=10, height=4, command=lambda: self.select_sites(self.sites_list))
        p = tk.Button(root, text="Select property to scan", wraplength=70, width=10, height=4, command=lambda: self.select_property(self.property_list))
        l = tk.Button(root, text="Leave", width=10, height=4, command=self.cancel)
        e = tk.Button(root, text="Scan/ Replace", wraplength=70, width=10, height=4, command=self.scan)
        r = tk.Button(root, text="Replace Interface", wraplength=70, width=10, height=4, command=lambda: self.toggle_widget(self.replace_bar))

        s.pack(in_=middle, side=tk.LEFT)
        p.pack(in_=middle, side=tk.LEFT)
        e.pack(in_=middle, side=tk.LEFT)
        r.pack(in_=middle, side=tk.LEFT)
        l.pack(in_=middle, side=tk.LEFT)

        # create the widgets for the bottom part of the GUI,
        # and lay them out
        global path
        path = tk.Label(root, text = "Select 'Scan' to read selected files and check the replace box to make replacements", width=35, height=5)
        path.pack(in_=top, side=tk.LEFT, fill=tk.BOTH, expand=True)
    
        self.root = root

        

#__________________________________________________________________________
# main 
with logger(LOGFILE) as log, logger(ERRORLOG) as error:

    # Load the Excel file
    try:
        os_list = pd.read_excel(EXCEL_FILE, SHEET_NAME)
    except Exception as e:
        log.login(f"error reading spreadsheets: {e}")

    #make sites list
    try:
        unique_Sites = sorted(set(os_list['siteLabel'].dropna()))
        log.login(f"sites list : {unique_Sites}")
    except Exception as e:
        log.login(f"error creating site list: {e}")

    #make property list
    try:
        with open('json_dat/pages.json', 'r') as json_file:
            pages = json.load(json_file)
            property_list = []
            for page, link in pages.items():
                property_list.append(page)
            sorted(set(property_list))
            log.login(f"property_list : {property_list}")
    except Exception as e:
        log.login(f"error creating property_list: {e}")

    
    
    except Exception as e:
        log.login(f"error getting IPs: {e}")

    #init mode
    Replace = False
    fetch_users()

    checkbox_var = False
    gui = GUI(unique_Sites, property_list)
    gui.root.mainloop()
    
